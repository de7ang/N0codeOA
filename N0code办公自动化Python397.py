# encoding: utf-8 2022-6-15
import pandas as pd
import numpy as np
import io
import os
import re
import time
import chardet
import zipfile
import shutil
import win32file
import datetime
from win32file import *
from contextlib import redirect_stderr
import openpyxl
from docx import Document
from pdf2docx import Converter
import pdfplumber
import pinyin


def atm_txt2excel():
    with open(Fi1eIN, 'rb') as openfile:
        enc0des = chardet.detect(openfile.readline())
        print("推测文件编码：", enc0des)
        enc0de = enc0des['encoding']
        if enc0de == "GB2312" or "ISO" in enc0de or enc0de is None:
            enc0de = "GB18030"
        print("确定文件编码：", enc0de)
    with open(Fi1eIN, 'r', encoding=enc0de) as openfile:
        text = openfile.read()
        text = text.replace("\n", "")
        text = text.split("日期:")
        df = pd.DataFrame(columns=["和谐掉了", "和谐掉了", "和谐掉了", "和谐掉了", "和谐掉了", "和谐掉了", "和谐掉了"])
        df_index = 1
        for texts in text:
            if texts:
                rmb_sn_findall = re.findall(r"(?<=LEVEL4:).*?\d{2}:\d{2}:\d{2} ", texts)
                trade_beg_time = re.findall(r"\d{2}:\d{2}:\d{2}", texts)
                if trade_beg_time:
                    trade_beg_time = trade_beg_time[0]
                else:
                    trade_beg_time = "未知"
                trade_date = re.findall(r"(\d{4}-\d{2}-\d{2})", texts)
                if trade_date:
                    trade_date = trade_date[0]
                else:
                    trade_date = "未知"
                atm_sn = re.findall(r"(?<=ATM:)\d{6}", texts)
                if atm_sn:
                    atm_sn = atm_sn[0]
                else:
                    atm_sn = "未知"
                bankcard_sn = re.findall(r"\d{19}", texts)
                if bankcard_sn:
                    bankcard_sn = bankcard_sn[0]
                else:
                    bankcard_sn = "未知"
                for rmb_sn_find in rmb_sn_findall:
                    trade_end_time = re.findall(r"\d{2}:\d{2}:\d{2}", rmb_sn_find)
                    if trade_end_time:
                        trade_end_time = trade_end_time[0]
                    else:
                        trade_end_time = "未知"
                    rmb_sn_find = rmb_sn_find.replace("  ", " ")
                    rmb_sn_find = rmb_sn_find.split(" ")
                    for rmb_sn in rmb_sn_find:
                        if len(rmb_sn) == 10:
                            trade_date_sign = trade_date.replace("-", "")
                            trade_beg_time_sign = trade_beg_time.replace(":", "")
                            trade_sign = f"{atm_sn}_{trade_date_sign}_{trade_beg_time_sign}"
                            lst = [trade_sign, rmb_sn, bankcard_sn, trade_date, atm_sn, trade_beg_time, trade_end_time]
                            df.loc[df_index] = lst
                            print(lst)
                            df_index += 1
    dftofile(df, Fi1eOUT)


def get_time(taskname):  # 装饰器：统计函数耗时
    def get_task(func):
        def inner(*arg, **kwarg):
            start_time = time.time()
            res = func(*arg, **kwarg)
            end_time = time.time()
            print(f'{taskname} 耗时：{end_time - start_time}秒')
            return res
        return inner
    return get_task


def idcard_hash18(eighteen_card):  # 计算身份证第18位校验值
    wi = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2, 1, ]
    vi = [1, 0, 'X', 9, 8, 7, 6, 5, 4, 3, 2, ]
    ai = []
    remaining = ''
    if len(eighteen_card) == 18:
        eighteen_card = eighteen_card[0:-1]
    if len(eighteen_card) == 17:
        s = 0
        for i in eighteen_card:
            ai.append(int(i))
        for i in range(17):
            s = s + wi[i] * ai[i]
        remaining = s % 11
    return 'X' if remaining == 2 else str(vi[remaining])


def idcard_add_18bit(df):  # apply函数：分析身份证 15转18 校验18、年龄、生日、性别
    idcard = str(df[C0L])
    if len(idcard) == 15:
        idcard = idcard[0:6] + '19' + idcard[6:15]
        idcard += idcard_hash18(idcard)
    elif len(idcard) == 18:
        idcard = idcard[0: -1]
        idcard += idcard_hash18(idcard)
    if len(idcard) == 18:
        age = datetime.date.today().year - int(idcard[6:10])
        birthday = idcard[6:14]
        if int(idcard[16]) % 2:  # 身份证号17位除以2 余1为奇数_男性_真 余0为偶数_女性_假
            sex = "男"
        else:
            sex = "女"
    else:
        age = birthday = sex = "空"
    return idcard, age, birthday, sex


def idcard_add_area(df):  # apply函数：分析身份证 地区
    idcard = str(df["_18位校验身份证"])
    if len(idcard) == 18:
        area = int(idcard[:6])
        if area in IDcardAERA.index:
            area = IDcardAERA.loc[area]
    else:
        area = "空"
    return area


def hanzi2pinyin(df):  # apply函数：获取指定列完整拼音、拼音首字母
    getpinyin = pinyin.get(str(df[C0L]), format='strip', delimiter=" ")
    getpy = pinyin.get_initial(str(df[C0L]), delimiter="")
    return getpinyin, getpy


def network2pinyin(df):  # apply函数：获取指定列完整拼音（d分析网点名@专用）
    getpinyin = pinyin.get(str(df["_网点精简名"]), format='strip', delimiter=" ")
    return getpinyin


def get_pinyin():  # f获取指定列拼音@ 指定单列
    df = file_read(Fi1eD1R, Fi1eFULL)
    cols = ["_完整拼音：" + C0L, "_拼音首字母：" + C0L]
    dfcols = pd.DataFrame(df.apply(hanzi2pinyin, axis=1).to_list(), columns=cols)
    for col in cols:
        df.insert(0, col, dfcols[col], allow_duplicates=True)
    dftofile(df, Fi1eOUT)


def idcard_analyse():  # 3分析处理 b分析身份证@ 指定单列
    global IDcardAERA
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        cols = ["_18位校验身份证", "_身份证年龄", "_身份证生日", "_身份证性别"]
        dfcols = pd.DataFrame(df.apply(idcard_add_18bit, axis=1).to_list(), columns=cols)
        for col in cols:
            df.insert(0, col, dfcols[col], allow_duplicates=True)
        IDcardAERA = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='身份证代码to地区', index_col="代码")
        df.insert(0, "_身份证地区", df.apply(idcard_add_area, axis=1)["地区"], allow_duplicates=True)
        dftofile(df, Fi1eOUT)
        IDcardAERA = pd.DataFrame()


def network_analyse(task):  # 3分析处理 c分析机构号@ d分析网点名@ 指定单列
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        id_name = pd.DataFrame()
        if task == "_机构号":
            df[task] = df[C0L].str[0:4]
            id_name = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='机构号to网点名', dtype=str)
        elif task == "_精简名拼音":
            redict = {"[^\u4e00-\u9fa5]": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", 和谐掉了": "",
                      "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": "", "和谐掉了": ""}
            df["_网点精简名"] = df[C0L].astype("str").replace(redict, regex=True)
            df.insert(0, "_精简名拼音", df.apply(network2pinyin, axis=1), allow_duplicates=True)
            id_name = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='网点名to机构号（拼音匹配）', dtype=str)
        dfresult = pd.merge(df, id_name, how="left", on=task)
        if task == "_机构号":
            dfresult = dfresult.drop(["_精简名", "_曾用旧名", "_曾用错名"], axis=1)
        elif task == "_精简名拼音":
            dfresult = dfresult.drop(["_精简名", "_精简名拼音"], axis=1)
        dftofile(dfresult, Fi1eOUT)


def excel_analyse():  # 3分析处理 e分析表格结构 可指定多列
    df = file_read(Fi1eD1R, Fi1eFULL, False)
    if not df.empty:
        col_value = ""
        if C0L:
            if not find_columns(df, C0L.split(",")):
                file_rename("列名不存在")
                return
            for colr in C0L.split(","):
                col_value += "    ******【" + colr + "】的列值及数量******\n" + str(df[colr].value_counts()) + "\n\n\n"
        cols = "    ******所有列的列名及第1行列值******\n"
        for col in df.columns:
            cols = cols + col + "\t：\t" + str(df[col][0]) + "\n"
        buf = io.StringIO()  # 创建一个StringIO，便于后续在内存中写入str
        df.info(buf=buf)  # 写入
        info = buf.getvalue()  # 读取
        redict = {"Column": "列名", "Dtype": "数据类型", "dtypes:": "数据类型计数：", "object": "文本", "int64": "整数",
                  "float64": "小数", "Non-Null Count": "非空行计数", "non-null": "非空行", "Data columns (total": "总列数：",
                  "columns):": "列", "RangeIndex:": "总行数：", "entries,": "行",
                  "<class 'pandas.core.frame.DataFrame'>": "\n\n    ******表格结构******"}
        for key, value in redict.items():
            info = info.replace(key, value)
        with open(Fi1eOUT + '.txt', 'a') as f:
            f.write(col_value + cols + info)


def word2excel():  # 4智能转换 word转excel
    document = Document(Fi1eIN)  # 读入文件
    tables = document.tables  # 获取文件中的表格集
    row_content = []
    for table in tables[:]:  # 记录表序号
        tb_list = []
        for row in table.rows[:]:  # 记录每个表的每一行存储于row中
            list0 = []
            for cell in row.cells[:]:  # 读一行中的所有单元格
                list0.append(cell.text)
            tb_list.append(list0)
        row_content.append(tb_list)
    if row_content:  # 源文件无表格，则返回1
        book = openpyxl.Workbook()  # 先创建一个工作簿
        del book["Sheet"]
        for s, tb in enumerate(row_content[:]):  # 读每个表数据
            sheet = book.create_sheet('Sheet' + str(s))  # 创建一个test_case的sheet表单
            for i, row in enumerate(tb[:]):  # 读每行数据
                for j, cell in enumerate(row[:]):  # 读每个单元格数据
                    sheet.cell(i + 1, j + 1, cell)
            time.sleep(1)
        book.save(Fi1eOUT + ".xlsx")


def pdf2office():  # 4智能转换 pdf转office
    cv = Converter(Fi1eIN)
    cv.convert(Fi1eOUT + ".docx", start=0, end=None)
    cv.close()
    pdf = pdfplumber.open(Fi1eIN)
    page = len(pdf.pages)
    print('总共有', page, '页')
    writer = pd.ExcelWriter(Fi1eOUT + ".xlsx", engine='xlsxwriter')
    for i in range(0, page):
        print('正在输出第', str(i + 1), '页表格')
        p0 = pdf.pages[i]
        try:
            table = p0.extract_table()
            df = pd.DataFrame(table[1:], columns=table[0])
            df.to_excel(writer, sheet_name='Sheet' + str(i), index=False)
        except Exception as e:
            print(e)
            pass
    writer.close()
    pdf.close()
    if os.path.getsize(Fi1eOUT + ".xlsx") < 5555:
        os.remove(Fi1eOUT + ".xlsx")


def ofd2txt():  # 4智能转换 ofd转txt
    pages = []
    with zipfile.ZipFile(Fi1eIN, 'r') as f:
        for file in f.namelist():
            f.extract(file, T3MP)
    for filedir, dirs, files in os.walk(T3MP + "Doc_0\\Pages"):
        for filename in files:
            if "Page_" in filedir:
                pages.append(filedir)
    pages.sort(key=lambda x: int(x[x.rfind("_"):].replace("_", "")))
    print(pages)
    total = ""
    for page in pages:
        xmlpath = page + "\\Content.xml"
        with open(xmlpath, 'r', encoding="utf-8") as openfile:
            text = openfile.read()
            text = re.sub(r"<.*?>", "", text)
            total += text
    if total:
        with open(f"{Fi1eOUT}.txt", 'w', encoding="utf-8") as openfile:
            openfile.writelines(total)


def split_df():  # 2筛选拆分 a拆分指定列@ 指定单列
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        groups = df.groupby(df[C0L])  # 仅支持单列
        if not os.path.exists(T3MP):
            os.mkdir(T3MP)
        for group in groups:
            group[1].to_excel(T3MP + str(group[0]) + '.xlsx', index=False)
        dirtozip(T3MP, Fi1eOUT)


def select_range(selectnum=True):  # 2筛选拆分 默认真_b筛选数值范围@#，可选假_c筛选时间范围@#（范围16770922-22620411） 指定单列
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        col = C0L
        if not selectnum:
            col = "_时间戳"
            df[col] = pd.to_datetime(df[C0L].astype(str))
        if (selectnum and ("str" not in str(type(df[col][0])))) or not selectnum:
            for range_num in R4NGE.split(','):
                print(f"筛选范围：{range_num}")
                filename = Fi1eN4ME.replace(R4NGE, range_num)
                outname = f"{Fi1eD1R}\\结果{T1ME}：{filename}"
                siftdf = pd.DataFrame()
                if "x-" in range_num:
                    if selectnum:
                        num = int(range_num.replace("x-", ""))
                    else:
                        num = pd.to_datetime(range_num.replace("x-", ""))
                    print(f"小于：{num}")
                    siftdf = df[df[col] < num]
                elif "-x" in range_num:
                    if selectnum:
                        num = int(range_num.replace("-x", ""))
                    else:
                        num = pd.to_datetime(range_num.replace("-x", ""))
                    print(f"大于：{num}")
                    siftdf = df[df[col] > num]
                elif "x=" in range_num:
                    if selectnum:
                        num = int(range_num.replace("x=", ""))
                    else:
                        num = pd.to_datetime(range_num.replace("x=", ""))
                    print(f"小于等于：{num}")
                    siftdf = df[df[col] <= num]
                elif "=x" in range_num:
                    if selectnum:
                        num = int(range_num.replace("=x", ""))
                    else:
                        num = pd.to_datetime(range_num.replace("=x", ""))
                    print(f"大于等于：{num}")
                    siftdf = df[df[col] >= num]
                elif "-" in range_num:
                    num = range_num.split('-')
                    if selectnum:
                        m1n = int(num[0])
                        m4x = int(num[1])
                    else:
                        m1n = pd.to_datetime(num[0])
                        m4x = pd.to_datetime(num[1])
                    print(f"大于：{m1n} 小于：{m4x}")
                    siftdf = df[(df[col] > m1n) & (df[col] < m4x)]
                elif "=" in range_num:
                    num = range_num.split('=')
                    if selectnum:
                        m1n = int(num[0])
                        m4x = int(num[1])
                    else:
                        m1n = pd.to_datetime(num[0])
                        m4x = pd.to_datetime(num[1])
                    print(f"大于等于：{m1n} 小于等于：{m4x}")
                    siftdf = df[(df[col] >= m1n) & (df[col] <= m4x)]
                dftofile(siftdf, outname)
        else:
            file_rename("筛选列含非数值")


def select_text():  # 2筛选拆分 d筛选指定文本@# 指定单列+参数
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        if "str" not in str(type(df[C0L][0])):
            df[C0L] = df[C0L].astype("str")
        for range_str in R4NGE.split(','):
            filename = Fi1eN4ME.replace(R4NGE, range_str)
            outname = f"{Fi1eD1R}\\结果{T1ME}：{filename}"
            range_str = range_str.replace("-", "|")
            if "半角？" in range_str:
                range_str = range_str.replace("半角？", "\\?")
            if "~" in range_str:
                range_str = range_str.replace("~", "")
                siftdf = df.loc[~df[C0L].str.contains(range_str)]
            else:
                siftdf = df.loc[df[C0L].str.contains(range_str)]
            dftofile(siftdf, outname)


def select_column():  # 2筛选拆分 g选取或删除列@@ 指定多列
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not find_columns(df, C0L.replace("~", "").split(',')):
        file_rename("列名不存在")
        return
    if not df.empty:
        if "~" in C0L:  # 删除指定列
            col = C0L.replace("~", "")
            siftdf = df.drop(col.split(','), axis=1)
            outname = f"{Fi1eD1R}\\结果{T1ME}已删除列：{Fi1eN4ME}"
        else:  # 选取指定列
            siftdf = df[C0L.split(',')]
            outname = f"{Fi1eD1R}\\结果{T1ME}已选取列：{Fi1eN4ME}"
        dftofile(siftdf, outname)


def select_query():  # 2筛选拆分 h自定义公式筛选 指定公式.txt
    unzip(T3MP)
    findfile = 0
    for filedir, dirs, files in os.walk(T3MP):
        for filename in files:
            if "公式.txt" in filename:
                filepath = os.path.join(filedir, "公式.txt")
                if os.path.getsize(filepath):  # 文件大小不等于0才执行
                    with open(filepath, 'rb') as openfile:
                        enc0des = chardet.detect(openfile.readline())
                        print("推测公式.txt编码", enc0des)
                        enc0de = enc0des['encoding']
                        if enc0de == "GB2312" or "ISO" in enc0de or enc0de is None:
                            enc0de = "GB18030"
                        print("确定公式.txt编码：" + enc0de)
                    with open(filepath, 'r', encoding=enc0de) as openfile:
                        range_querys = openfile.readlines()
                        print("全部自定义公式：", range_querys)
                    findfile = 1
                else:
                    file_rename("公式.txt没有内容")
                    return
    if findfile:
        for filedir, dirs, files in os.walk(T3MP):
            for filename in files:
                if "公式.txt" not in filename:
                    df = file_read(filedir, filename)
                    if not df.empty:
                        i = 0
                        for range_query in range_querys:
                            print(f"第{i}行公式：{range_query}")
                            i += 1
                            siftdf = df.query(range_query)
                            outname = f"{Fi1eD1R}\\结果{T1ME}第{i}行公式：{filename}"
                            dftofile(siftdf, outname)
    else:
        file_rename("未指定公式")
        return


def del_duplicates():  # 2筛选拆分 2e删除重复值@@# 指定多列+参数
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        col = C0L.split(',')
        if not find_columns(df, col):
            file_rename("列名不存在")
            return
        if R4NGE == "0":
            df.drop_duplicates(subset=col, keep=False, inplace=True)
            print("False： 删除所有重复值")
        elif R4NGE == "1":
            df.drop_duplicates(subset=col, keep="last", inplace=True)
            print("last： 保留最后一次出现的重复值")
        else:
            df.drop_duplicates(subset=col, keep="first", inplace=True)
            print("first： 保留第一次出现的重复值")
        dftofile(df, Fi1eOUT)


def select_duplicates():  # 2筛选拆分 2f选取重复值@@# 指定多列+参数
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        col = C0L.split(',')
        if not find_columns(df, col):
            file_rename("列名不存在")
            return
        if R4NGE == "0":
            df = df[df.duplicated(subset=col, keep="first")]
            print("first： 除了第一次出现外，其余相同的被标记为重复")
        elif R4NGE == "1":
            df = df[df.duplicated(subset=col, keep="last")]
            print("last：除了最后一次出现外，其余相同的被标记为重复")
        else:
            df = df[df.duplicated(subset=col, keep=False)]
            print("False：即所有相同的都被标记为重复")
        dftofile(df, Fi1eOUT)


def append_df():  # 1匹配合并 f多表对齐同名列_纵向堆叠
    unzip(T3MP)
    df = pd.DataFrame()
    for filedir, dirs, files in os.walk(T3MP):
        for filename in files:
            unzipdf = file_read(filedir, filename)
            if not unzipdf.empty:
                df = df.append(unzipdf)
    dftofile(df, Fi1eOUT)


def merge_df(h0w, notboth=False):  # 1匹配合并 两表横向@@ 指定多列
    dfleft = pd.DataFrame()
    dfright = pd.DataFrame()
    c0lleft = C0L.split(',')
    c0lright = C0L.split(',')
    if ".zip" in Fi1eEXT:
        unzip(T3MP)
        for filedir, dirs, files in os.walk(T3MP):
            for filename in files:
                col0 = re.findall(r"\[.*]", filename)
                if "左表" in filename:
                    if col0:
                        c0lleft = re.sub(r"[][]", "", col0[0])  # 提取中括号里的列名
                        c0lleft = c0lleft.split(',')
                    dfleft = file_read(filedir, filename)
                if "右表" in filename:
                    if col0:
                        c0lright = re.sub(r"[][]", "", col0[0])  # 提取中括号里的列名
                        c0lright = c0lright.split(',')
                    dfright = file_read(filedir, filename)
        if dfleft.empty:
            file_rename("未指定左表")
            return
        if dfright.empty:
            file_rename("未指定右表")
            return
        if not c0lleft[0]:
            file_rename("未指定左表列名")
            return
        if not c0lright[0]:
            file_rename("未指定右表列名")
            return
    else:
        if C0L:
            dfleft = pd.read_excel(Fi1eIN, sheet_name='左表', keep_default_na=False)
            dfright = pd.read_excel(Fi1eIN, sheet_name='右表', keep_default_na=False)
            # 预处理：所有数值列11位及以下转为文本，所有文本列去除TAB空白
            dfleft = file_read_transform(dfleft)
            dfright = file_read_transform(dfright)
        else:
            file_rename("未指定列名")
            return
    if '_本行来自' in dfleft.columns:
        dfleft = dfleft.drop('_本行来自', axis=1)
    if '_本行来自' in dfright.columns:
        dfright = dfright.drop('_本行来自', axis=1)
    print(f"左列名：【{c0lleft}】 右列名：【{c0lright}】")
    if not find_columns(dfleft, c0lleft):
        file_rename("列名不存在")
        return
    if not find_columns(dfright, c0lright):
        file_rename("列名不存在")
        return
    df = pd.merge(dfleft, dfright, how=h0w, left_on=c0lleft, right_on=c0lright, suffixes=('_来自左表', '_来自右表'),
                  indicator='_本行来自')
    df['_本行来自'].replace({'both': '两表', 'left_only': '左表', 'right_only': '右表'}, inplace=True)
    if notboth:
        df = df[df['_本行来自'] != '两表']
    dftofile(df, Fi1eOUT)


def merge_db():  # 5 数据库比对
    if not DataBa5e.empty:
        dfright = file_read(Fi1eD1R, Fi1eFULL)
        if '_本行来自' in dfright.columns:
            dfright = dfright.drop('_本行来自', axis=1)
        if C0L:
            right_col = C0L.split(",")
        else:
            right_col = DataBa5eC0L
        print("数据库指定列：", DataBa5eC0L)
        print("传入表指定列：", right_col)
        dbhow = ""
        if "a输出交集" in Fi1eD1R:
            dbhow = "inner"
        elif "b输出原表" in Fi1eD1R:
            dbhow = "right"
        df = pd.merge(DataBa5e, dfright, how=dbhow, left_on=DataBa5eC0L, right_on=right_col,
                      suffixes=('_来自数据库', '_来自原表'), indicator='_本行来自')
        df['_本行来自'].replace({'both': '两表', 'left_only': '数据库', 'right_only': '原表'}, inplace=True)
        if df.shape[0] > 0:
            dftofile(df, Fi1eOUT)
        else:
            file_rename("", "原件成功_没有交集")
            print("无")
    else:
        file_rename("请更新数据库")


def find_columns(df, cols):  # 查找列名cols（列表类型）是否存在
    finds = 0
    for col in cols:
        find = False
        for dfcol in df.columns:
            if col == dfcol:
                find = True
        if find:
            finds += 1
    if finds == len(cols):
        return True
    else:
        return False


def excelpivottable_1col_backup():  # 3分析处理 a生成数据透视表@ 指定单列（备用方案）
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            file_rename("列名不存在")
            return
        list_num_time = list(df.select_dtypes(include=["int", "int64", "float", "float64", "datetime", "datetime64"]))
        print("类型为数值和时间的列：", list_num_time)
        if list_num_time:
            col = C0L
            list_num_time = [i for i in list_num_time if i not in col]
            print("去掉指定列后的列：", list_num_time)
            dflen = pd.pivot_table(df, index=[col], values=list_num_time[0], aggfunc=len)
            dflen.columns = ["计数"]
            dfsum = pd.pivot_table(df, index=[col], values=list_num_time, aggfunc=np.sum)
            dfsum.columns = "_求和：" + dfsum.columns
            dfmean = pd.pivot_table(df, index=[col], values=list_num_time, aggfunc=np.mean)
            dfmean.columns = "_平均值：" + dfmean.columns
            dfmax = pd.pivot_table(df, index=[col], values=list_num_time, aggfunc=max)
            dfmax.columns = "_最大值：" + dfmax.columns
            dfmin = pd.pivot_table(df, index=[col], values=list_num_time, aggfunc=min)
            dfmin.columns = "_最小值：" + dfmin.columns
            df = pd.concat([dflen, dfsum, dfmean, dfmax, dfmin], axis=1)
            df.reset_index(level=0, inplace=True)
            dftofile(df, Fi1eOUT)


def excelpivottable():  # 3分析处理 a生成数据透视表@ 指定多列
    df = file_read(Fi1eD1R, Fi1eFULL)
    if not df.empty:
        if not find_columns(df, C0L.split(",")):
            file_rename("列名不存在")
            return
        list_num_time = list(df.select_dtypes(include=["int", "int64", "float", "float64", "datetime", "datetime64"]))
        print("类型为数值和时间的列：", list_num_time)
        if list_num_time:
            col = C0L.split(",")
            list_num_time = [i for i in list_num_time if i not in col]
            print("去掉指定列后的列：", list_num_time)
            dfresult = df.groupby(col, as_index=False)[list_num_time[0]].agg(len)
            dfresult.rename(columns={list_num_time[0]: "_计数"}, inplace=True)
            for key, value in {np.sum: "_求和：", np.mean: "_平均值：", max: "_最大值：", min: "_最小值："}.items():
                dfpivot = df.groupby(col, as_index=False)[list_num_time].agg(key)
                dfpivot = dfpivot.drop(col, axis=1)
                for i in list_num_time:
                    dfpivot.rename(columns={i: value + i}, inplace=True)
                dfresult = pd.concat([dfresult, dfpivot], axis=1)
            dftofile(dfresult, Fi1eOUT)


def unzip(temp):  # 解压文件
    with zipfile.ZipFile(file=Fi1eIN, mode='r') as zf:
        if not os.path.exists(temp):
            os.mkdir(temp)
        for oldn4me in zf.namelist():
            file_size = zf.getinfo(oldn4me).file_size
            newn4me = oldn4me.encode('cp437').decode('gbk')
            p4th = os.path.join(temp, newn4me)
            if file_size > 0:
                with open(file=p4th, mode='wb') as f:
                    f.write(zf.read(oldn4me))
            else:
                os.mkdir(p4th)


def dirtozip(dirin, fileout):  # df写入压缩文件
    zipout = fileout.replace(RESU1T, Fi1eD1R) + ".zip"
    with zipfile.ZipFile(zipout, 'w', zipfile.ZIP_DEFLATED, allowZip64=True) as zf:
        for dirpath, dirnames, filenames in os.walk(dirin):
            for filename in filenames:
                zf.write(filename=dirpath + filename, arcname=filename)


def dftofile(df, fileout):  # df写入文件 Excel使用zip64
    print("写入文件总行数：" + str(df.shape[0]))
    if (df.shape[0] >= 1048576) or ("tocsv" in Fi1eN4ME):
        if not os.path.exists(RESU1T):
            os.mkdir(RESU1T)
        result = fileout.replace(Fi1eD1R, RESU1T)
        df.to_csv(result + '.csv', index=False)
        dirtozip(RESU1T, result)
    else:
        writer = pd.ExcelWriter(fileout + ".xlsx", engine='xlsxwriter')
        df.to_excel(writer, index=False, header=True)
        writer.book.use_zip64()
        writer.save()
    print("【写入文件】成功：" + fileout)


def is_open():  # 判断文件是否打开
    try:
        vhandle = win32file.CreateFile(Fi1eIN, GENERIC_READ, 0, None, OPEN_EXISTING, FILE_ATTRIBUTE_NORMAL, None)
        if int(vhandle) == INVALID_HANDLE_VALUE:
            print("# 文件被占用中")
            return True  # file is already open
        win32file.CloseHandle(vhandle)
    except Exception as e:
        print(e)
        return True


def file_rename(fail, result="原件失败"):  # 文件重命名
    rename = f"{Fi1eD1R}\\{result}{T1ME}{fail}：{Fi1eFULL}"
    if os.path.exists(Fi1eIN):
        os.rename(Fi1eIN, rename)
        print(f"原始文件名已改为：{rename}")


@get_time("读取文件")
def file_read(filedir, filename, transform=True):  # 根据文件.扩展名读取文件，transform默认为真：要做预处理
    filepath = os.path.join(filedir, filename)  # 读取文件路径
    if "." in filename:
        fileext = filename[filename.rfind("."):]  # 文件.扩展名
    else:
        fileext = ""
    filename = filename.replace(fileext, "")  # 去掉文件.扩展名
    fileext = fileext.lower()
    df = pd.DataFrame()
    if os.path.getsize(filepath):  # 文件大小不等于0才执行
        if ".xls" in fileext:
            df = pd.read_excel(filepath, keep_default_na=False)
        else:
            # 判断csv文件编码及分隔符
            with open(filepath, 'rb') as openfile:
                enc0des = chardet.detect(openfile.readline())
                print("推测文件编码：", enc0des)
                enc0de = enc0des['encoding']
                if enc0de == "GB2312" or "ISO" in enc0de or enc0de is None:
                    enc0de = "GB18030"
                print("确定文件编码：", enc0de)
            with open(filepath, 'r', encoding=enc0de) as openfile:
                text3 = openfile.readline() + openfile.readline() + openfile.readline()
                print("文件前3行内容：", text3)
                text3 = re.sub("[^,;|.\t]", "", text3)  # 选取分隔符：【,】【;】【|】【.】【   】
                print("文件前3行分隔符：", text3)
                textsep = textmax(text3)
                print("判定分隔符为：【" + textsep + "】")
            # 处理跳行警告
            with open("Log.txt", 'w') as openfile:
                with redirect_stderr(openfile):
                    df = pd.read_csv(filepath, encoding=enc0de, sep=textsep, engine='python', error_bad_lines=False,
                                     keep_default_na=False)  # skipinitialspace =  忽略分隔符后的空白
            with open("Log.txt", 'r') as openfile:
                textlist = openfile.readlines()
                filetext = []
                for text in textlist:
                    if "Skipping line" in text:
                        text = text.replace("Skipping line", "最终结果不含原文件的第")
                        text = text.replace(":", "行，请检查，具体原因：")
                        filetext.append(text)
            if filetext:
                with open(f"{Fi1eD1R}\\跳行警告{T1ME}：{filename}.txt",
                          'w') as openfile:
                    openfile.writelines(filetext)
            if os.path.exists("Log.txt"):
                os.remove("Log.txt")
                time.sleep(1)
        print(f"【读取文件】成功：{filename}，总行数{df.shape[0]}")
        if transform:  # transform默认为真：做预处理：所有数值列11位及以下转为文本，所有文本列去除TAB空白
            df = file_read_transform(df)
    else:
        file_rename("文件为空")
        return df
    return df


def file_read_transform(df):  # 预处理：所有数值列11位及以下转为文本，所有文本列去除TAB空白
    # 所有数值列11位及以下转为文本
    list_num = list(df.select_dtypes(include=["int", "int64", "uint64", "float", "float64"]))
    print("所有类型为数值的列：", list_num)
    for i in list_num:
        # print(max(df[i].astype("str").apply(len)))  # 每列最大位数
        if not max(df[i].astype("str").apply(len)) <= 11:
            df[i] = df[i].astype("str")
    list_num = list(df.select_dtypes(include=["int", "int64", "uint64", "float", "float64"]))
    print(f"11位及以下的数值列：{list_num}")
    # 所有文本列去除TAB空白
    list_str = list(df.select_dtypes(include=["object"]))
    for i in list_str:
        df[i].replace("\t", "", regex=True, inplace=True)  # 正则去除所有TAB空白
    return df


def textmax(a):  # 计算字符串出现最多的符号
    t = ""
    k = 0
    y = []
    for q in a:
        s = a.count(q)
        y.append(s)
        t = max(y)
    for i in range(len(a)):
        if t != y[i]:
            continue
        k = a[i]
    print("出现最多的符号：【" + k + "】，" + str(t) + "次")
    return k


def scanfile(scandir):  # 扫描文件
    global Fi1eIN, Fi1eN4ME, Fi1eD1R, Fi1eOUT, Fi1eEXT, C0L, R4NGE, Fi1eFULL, T1ME, DataBa5e, DataBa5eC0L
    for Fi1eD1R, dirs, files in os.walk(scandir):
        files = [fi for fi in files if not fi.startswith(("~$", "结果", "失败原因", "跳行警告", "原件"))]
        files = [fi for fi in files if not fi.endswith((".bat", ".cmd"))]
        for Fi1eFULL in files:
            Fi1eIN = os.path.join(Fi1eD1R, Fi1eFULL)
            if not is_open():
                C0L = ""  # 列名
                R4NGE = ""  # 范围
                if "." in Fi1eFULL:
                    Fi1eEXT = Fi1eFULL[Fi1eFULL.rfind("."):]  # 文件.扩展名
                else:
                    Fi1eEXT = ""
                Fi1eN4ME = Fi1eFULL.replace(Fi1eEXT, "")  # 去掉文件.扩展名
                Fi1eEXT = Fi1eEXT.lower()
                print("小写扩展名", Fi1eEXT)
                T1ME = datetime.datetime.now().strftime('%H%M%S')
                Fi1eOUT = f"{Fi1eD1R}\\结果{T1ME}：{Fi1eN4ME}"
                col0 = re.findall(r"\[.*]", Fi1eN4ME)  # 提取中括号里的列名
                range0 = re.findall(r"{.*}", Fi1eN4ME)  # 提取大括号里的范围
                if col0:
                    C0L = re.sub(r"[][]", "", col0[0])  # 提取中括号里的列名
                if range0:
                    R4NGE = re.sub(r"[{}]", "", range0[0])  # 提取小括号里的范围
                print("────────────────────────────────────────────────────────────────────────────────────")
                print(f"【发现文件】全名：【{Fi1eN4ME}】 扩展名：【{Fi1eEXT}】 列名：【{C0L}】 范围：【{R4NGE}】")
                try:
                    start_times = time.time()
                    if "1匹配合并" in Fi1eD1R:
                        if "a两表横向并集_输出全部@" in Fi1eD1R:
                            print("【执行】1匹配合并 a两表横向并集_输出全部@")
                            merge_df("outer")
                        elif "b两表横向交集_输出相同@" in Fi1eD1R:
                            print("【执行】1匹配合并 b两表横向交集_输出相同@")
                            merge_df("inner")
                        elif "c两表横向反集_输出差异@" in Fi1eD1R:
                            print("【执行】1匹配合并 c_两表横向合并_反集")
                            merge_df("outer", True)
                        elif "d两表横向左集_输出左表@" in Fi1eD1R:
                            print("【执行】1匹配合并 d两表横向左集_输出左表@")
                            merge_df("left")
                        elif "e两表横向右集_输出右表@" in Fi1eD1R:
                            print("【执行】1匹配合并 e两表横向右集_输出右表@")
                            merge_df("right")
                        elif "f多表对齐同名列_纵向堆叠" in Fi1eD1R:
                            print("【执行】1匹配合并 f多表对齐同名列_纵向堆叠")
                            append_df()
                    elif "h自定义公式筛选" in Fi1eD1R:
                        print("【执行】2筛选拆分 h自定义公式筛选")
                        select_query()
                    elif "zip" in Fi1eEXT:
                        unzip(Fi1eD1R)
                        if os.path.exists(Fi1eIN):
                            os.remove(Fi1eIN)
                            print("  删除文件：" + Fi1eIN)
                    elif "2筛选拆分" in Fi1eD1R:
                        if ("a拆分指定列@" in Fi1eD1R) and C0L:
                            print("【执行】2筛选拆分 a拆分指定列@")
                            split_df()
                        elif ("b筛选数值范围@#" in Fi1eD1R) and C0L and R4NGE:
                            print("【执行】2筛选拆分 b筛选数值范围@#")
                            select_range(True)  # 真_筛选数值
                        elif ("c筛选时间范围@#" in Fi1eD1R) and C0L and R4NGE:
                            print("【执行】2筛选拆分 c筛选时间范围@#")
                            select_range(False)  # 真_筛选时间
                        elif ("d筛选指定文本@#" in Fi1eD1R) and C0L and R4NGE:
                            print("【执行】2筛选拆分 d筛选指定文本@#")
                            select_text()
                        elif ("e删除重复值@@#" in Fi1eD1R) and C0L:
                            print("【执行】2筛选拆分 e删除重复值@@#")
                            del_duplicates()
                        elif ("f选取重复值@@#" in Fi1eD1R) and C0L:
                            print("【执行】2筛选拆分 f选取重复值@@#")
                            select_duplicates()
                        elif ("g选取或删除列@@" in Fi1eD1R) and C0L:
                            print("【执行】2筛选拆分 g选取或删除列@@")
                            select_column()
                        else:
                            file_rename("未指定列名或参数")
                    elif "3分析处理" in Fi1eD1R:
                        if ("a生成数据透视表@" in Fi1eD1R) and C0L:
                            print("【执行】3分析处理 a生成数据透视表@")
                            excelpivottable()
                        elif ("b分析身份证@" in Fi1eD1R) and C0L:
                            print("【执行】3分析处理 b分析身份证@")
                            idcard_analyse()
                        elif ("c分析机构号@" in Fi1eD1R) and C0L:
                            print("【执行】3分析处理 c分析机构号@")
                            network_analyse("_机构号")
                        elif ("d分析网点名@" in Fi1eD1R) and C0L:
                            print("【执行】3分析处理 d分析网点名@")
                            network_analyse("_精简名拼音")
                        elif "e分析表格结构" in Fi1eD1R:
                            print("【执行】3分析处理 e分析表格结构")
                            excel_analyse()
                        elif ("f获取指定列拼音@" in Fi1eD1R) and C0L:
                            print("【执行】3分析处理 f获取指定列拼音@")
                            get_pinyin()
                        elif "g冠字号ATM流水转表格" in Fi1eD1R:
                            atm_txt2excel()
                        else:
                            file_rename("未指定列名或参数")
                    elif "4智能转换" in Fi1eD1R:
                        print("【执行】4智能转换")
                        if "pdf" in Fi1eEXT:
                            pdf2office()
                        elif "docx" in Fi1eEXT:
                            word2excel()
                        elif "doc" in Fi1eEXT:
                            file_rename("不支持doc，请先转换为docx")
                        elif ".ofd" in Fi1eEXT:
                            ofd2txt()
                        else:
                            df = file_read(Fi1eD1R, Fi1eFULL)
                            if not df.empty:
                                dftofile(df, Fi1eOUT)
                    elif ("数据库存放" in Fi1eD1R) or ("数据库更新" in Fi1eD1R):
                        if C0L:
                            DataBa5e = file_read(Fi1eD1R, Fi1eFULL)
                            DataBa5eC0L = C0L.split(',')
                            print("【数据库更新成功】\n", DataBa5e)
                            if "数据库存放" in Fi1eD1R:
                                return
                        else:
                            file_rename("未指定列名")
                            return
                    elif ("5数据库比对" in Fi1eD1R) and ("数据库更新" not in Fi1eD1R):
                        merge_db()
                    file_rename("", "原件成功")
                    end_times = time.time()
                    print("【执行成功】")
                    print(f'任务总耗时：{end_times - start_times}秒')
                except Exception as e:
                    file_rename("详见txt")
                    filepath = f"{Fi1eD1R}\\失败原因{T1ME}：{Fi1eN4ME}.txt"
                    with open(filepath, 'w') as openfile:
                        openfile.write(str(e))
                    print(e, "【执行失败】原因详见txt")
                finally:
                    if "数据库存放" not in Fi1eD1R:
                        file_rename("", "原件")
                    deldirs()
                    print("────────────────────────────────────────────────────────────────────────────────────")


def deldirs():  # 删除目录
    if os.path.exists(T3MP):
        shutil.rmtree(T3MP)
        print("删除临时目录：" + T3MP)
    if os.path.exists(RESU1T):
        shutil.rmtree(RESU1T)
        print("删除结果目录：" + RESU1T)


def makedirs():  # 创建目录
    if not os.path.exists(D1R):
        os.mkdir(D1R)
    for i in D1R_0:
        md = os.path.join(D1R, i)
        if not os.path.exists(md):
            os.mkdir(md)
    makedir(D1R_1, "1匹配合并")
    makedir(D1R_2, "2筛选拆分")
    makedir(D1R_3, "3分析处理")
    makedir(D1R_5, "5数据库比对")


def makedir(dir_x, dir_0):  # 创建子目录
    for i in dir_x:
        md = os.path.join(D1R, dir_0, i)
        if not os.path.exists(md):
            os.mkdir(md)


T3MP = os.getcwd() + "\\TEMP\\"
RESU1T = os.getcwd() + "\\结果\\"
DBsave = os.getcwd() + "\\数据库存放\\"
D1R = os.getcwd() + "\\功能\\"  # 功能根目录

D1R_0 = ['1匹配合并', '2筛选拆分', '3分析处理', '4智能转换', '5数据库比对']  # 功能主目录
D1R_1 = ['a两表横向并集_输出全部@@', 'b两表横向交集_输出相同@@', 'c两表横向反集_输出差异@@',
         'd两表横向左集_输出左表@@', 'e两表横向右集_输出右表@@', 'f多表对齐同名列_纵向堆叠']
D1R_2 = ['a拆分指定列@', 'b筛选数值范围@#', 'c筛选时间范围@#', 'd筛选指定文本@#',
         'e删除重复值@@#', 'f选取重复值@@#', 'g选取或删除列@@', 'h自定义公式筛选']
D1R_3 = ['a生成数据透视表@', 'b分析身份证@', 'c分析机构号@', 'd分析网点名@', 'e分析表格结构', 'f获取指定列拼音@', 'g冠字号ATM流水转表格']
D1R_5 = ['数据库更新', 'a输出交集', 'b输出原表']
Fi1eD1R = ""  # 输入文件目录路径
Fi1eFULL = ""  # 输入文件 名字.扩展名
Fi1eN4ME = ""  # 输入文件 名字
Fi1eEXT = ""  # 输入文件 .扩展名
Fi1eIN = ""  # 输入文件路径
Fi1eOUT = ""  # 输出文件路径
C0L = ""  # 列名
R4NGE = ""  # 范围
T1ME = ""  # 时分秒
IDcardAERA = pd.DataFrame()  # _机构号网点名_身份证代码地区_转换表.xlsx sheet_name='身份证代码to地区'
DataBa5e = pd.DataFrame()  # 待比对数据库
DataBa5eC0L = ""  # 待比对数据库列名
if __name__ == "__main__":
    print("────────────────────────────────────────────────────────────────────────────────────")
    print("                        程序运行中，请将文件放入功能中的对应目录")
    print("────────────────────────────────────────────────────────────────────────────────────")
    makedirs()
    if os.path.exists(DBsave):
        scanfile(DBsave)
    else:
        os.mkdir(DBsave)
    while True:
        time.sleep(1)
        deldirs()
        try:
            scanfile(D1R)
        except RuntimeError as ec:
            pass
