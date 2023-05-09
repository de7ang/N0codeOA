# encoding: utf-8
# 2023-4-27
import os
import re
import pinyin
import datetime
import pdfplumber
import numpy as np
import pandas as pd
from io import StringIO
from docx import Document
from shutil import rmtree
from chardet import detect
from time import sleep, time
from openpyxl import Workbook
from pdf2docx import Converter
from contextlib import redirect_stderr
from zipfile import ZipFile


def get_date():  # 3分析处理 h八位列转日期格式
    global 八位日期列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        for 八位日期列 in df.columns:
            if max(df[八位日期列].astype("str").apply(len)) == 8:
                df.insert(0, "_转日期格式：" + 八位日期列, df.apply(col2date, axis=1), allow_duplicates=True)
    dftofile(df)


def get_time(taskname):  # 装饰器：统计函数耗时
    def get_task(func):
        def inner(*arg, **kwarg):
            start_time = time()
            res = func(*arg, **kwarg)
            end_time = time()
            print(f'{taskname} 耗时：{end_time - start_time}秒')
            return res
        return inner
    return get_task


def idcard_hash18(eighteen_card):  # 计算身份证第18位校验值
    wi = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2, 1, ]
    vi = [1, 0, 'X', 9, 8, 7, 6, 5, 4, 3, 2, ]
    ai = []
    remaining = ''
    if len(eighteen_card) == 18:
        eighteen_card = eighteen_card[0:-1]
    if len(eighteen_card) == 17:
        s = 0
        for i in eighteen_card:
            ai.append(int(i))
        for i in range(17):
            s = s + wi[i] * ai[i]
        remaining = s % 11
    return 'X' if remaining == 2 else str(vi[remaining])


def idcard_add_18bit(df):  # apply函数：分析身份证 15转18 校验18、年龄、生日、性别
    idcard = str(df[C0L])
    idcardsource = "非身份证保持原样"
    if len(idcard) == 15:
        idcard = idcard[0:6] + '19' + idcard[6:15]
        idcard += idcard_hash18(idcard)
        idcardsource = "经15位身份证转换"
    elif len(idcard) == 18:
        idcard_hash = idcard[0: -1]
        idcard_hash += idcard_hash18(idcard)
        if idcard_hash != idcard:
            idcard = idcard_hash
            idcardsource = "错误18位身份证重新校验"
        else:
            idcardsource = "正确身份证保持原样"
    if len(idcard) == 18:
        age = datetime.date.today().year - int(idcard[6:10])
        birthday = idcard[6:14]
        if int(idcard[16]) % 2:  # 身份证号17位除以2 余1为奇数_男性_真 余0为偶数_女性_假
            sex = "男"
        else:
            sex = "女"
        area = int(idcard[:6])
        if area in IDcardAERA.index:
            area = IDcardAERA.loc[area][0]
        else:
            area = "未查询到"
    else:
        age = birthday = sex = area = "空"
    return idcardsource, idcard, age, birthday, sex, area


def hanzi2pinyin(df):  # apply函数：获取指定列完整拼音、拼音首字母
    getpinyin = pinyin.get(str(df[C0L]), format='strip', delimiter=" ")
    getpy = pinyin.get_initial(str(df[C0L]), delimiter="")
    return getpinyin, getpy


def network2pinyin(df):  # apply函数：获取指定列完整拼音（d分析网点名@专用）
    getpinyin = pinyin.get(str(df["_网点精简名"]), format='strip', delimiter=" ")
    return getpinyin


def isvaliddate(date):  # 校验日期是否合法
    try:
        date = date.split("-")
        year = int(date[0])
        month = int(date[1])
        day = int(date[2])
        if month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
            if day > 31:
                return False
            else:
                return True
        if month == 4 or month == 6 or month == 9 or month == 11:
            if day > 30:
                return False
            else:
                return True
        if month == 2:
            if year % 4 == 0 & year % 400 != 0:
                if day > 29:
                    return False
                else:
                    return True
            else:
                if day > 28:
                    return False
                else:
                    return True
    except Exception as e:
        print(e)
        return False


def col2date(df):  # apply函数：获取指定列日期格式
    date = str(df[八位日期列])
    if len(date) == 8:
        date = date[0:4] + "-" + date[4:6] + "-" + date[6:8]
        if isvaliddate(date):
            date = datetime.datetime.strptime(date, "%Y-%m-%d")
            date = datetime.datetime.date(date)
        else:
            date = "不含标准日期"
    else:
        date = "不含标准日期"
    return date


def get_pinyin():  # f获取指定列拼音@ 指定单列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    cols = ["_完整拼音：" + C0L, "_拼音首字母：" + C0L]
    dfcols = pd.DataFrame(df.apply(hanzi2pinyin, axis=1).to_list(), columns=cols)
    for col in cols:
        df.insert(0, col, dfcols[col], allow_duplicates=True)
    dftofile(df)


def idcard_analyse():  # 3分析处理 b分析身份证@ 指定单列
    global IDcardAERA
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            print("表格中不存在指定的列名，请核实")
            return
        IDcardAERA = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='身份证代码to地区', index_col="代码")
        cols = ["_18位校验身份证来源", "_18位校验身份证", "_身份证年龄", "_身份证生日", "_身份证性别", "_身份证地区"]
        dfcols = pd.DataFrame(df.apply(idcard_add_18bit, axis=1).to_list(), columns=cols)
        for col in cols:
            df.insert(0, col, dfcols[col], allow_duplicates=True)
        dftofile(df)
        IDcardAERA = pd.DataFrame()


def network_analyse(task):  # 3分析处理 c分析机构号@ d分析网点名@ 指定单列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            print("表格中不存在指定的列名，请核实")
            return
        id_name = pd.DataFrame()
        if task == "_机构号":
            df[task] = df[C0L].str[0:4]
            id_name = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='机构号to网点名', dtype=str)
        elif task == "_精简名拼音":
            redict = {"[^\u4e00-\u9fa5]": "", "四川省": "", "四川": "", "资阳市": "", "资阳": "", "农村": "", "商业": "",
                      "银行": "", "股份": "", "有限": "", "公司": "", "农商行": "", "农商": "", "支行": ""}
            df["_网点精简名"] = df[C0L].astype("str").replace(redict, regex=True)
            df.insert(0, "_精简名拼音", df.apply(network2pinyin, axis=1), allow_duplicates=True)
            id_name = pd.read_excel("_机构号网点名_身份证代码地区_转换表.xlsx", sheet_name='网点名to机构号（拼音匹配）', dtype=str)
        dfresult = pd.merge(df, id_name, how="left", on=task)
        if task == "_机构号":
            dfresult = dfresult.drop(["_精简名", "_曾用旧名", "_曾用错名"], axis=1)
        elif task == "_精简名拼音":
            dfresult = dfresult.drop(["_精简名", "_精简名拼音"], axis=1)
        dftofile(dfresult)


def excel_analyse():  # 3分析处理 e分析表格结构 可指定多列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径), False)
    if not df.empty:
        col_values = ""
        if C0L:
            if not find_columns(df, C0L.split(",")):
                print("表格中不存在指定的列名，请核实")
                return
            for colr in C0L.split(","):
                col_value = ""
                for k, v in dict(df[colr].value_counts()).items():
                    col_value += (str(k) + '\t：\t' + str(v) + '\n')
                col_values += "    ******【" + colr + "】的列值及数量******\n" + col_value + "\n\n\n"
        cols = "    ******所有列的列名及第1行列值******\n"
        for col in df.columns:
            cols = cols + col + "\t：\t" + str(df[col][0]) + "\n"
        buf = StringIO()  # 创建一个StringIO，便于后续在内存中写入str
        df.info(buf=buf)  # 写入
        info = buf.getvalue()  # 读取
        redict = {"Column": "列名", "Dtype": "数据类型", "dtypes:": "数据类型计数：", "object": "文本", "int64": "整数",
                  "float64": "小数", "Non-Null Count": "非空行计数", "non-null": "非空行", "Data columns (total": "总列数：",
                  "columns):": "列", "RangeIndex:": "总行数：", "entries,": "行",
                  "<class 'pandas.core.frame.DataFrame'>": "\n\n    ******表格结构******"}
        for key, value in redict.items():
            info = info.replace(key, value)
        dfoutname = os.path.join(D0R, OutN4ME)
        with open(dfoutname, 'a') as f:
            f.write(col_values + cols + info)


def word2excel():  # 4智能转换 word转excel
    document = Document(左表路径)  # 读入文件
    tables = document.tables  # 获取文件中的表格集
    row_content = []
    dfoutname = os.path.join(D0R, OutN4ME).replace(".txt", ".xlsx")
    for table in tables[:]:  # 记录表序号
        tb_list = []
        for row in table.rows[:]:  # 记录每个表的每一行存储于row中
            list0 = []
            for cell in row.cells[:]:  # 读一行中的所有单元格
                list0.append(cell.text)
            tb_list.append(list0)
        row_content.append(tb_list)
    if row_content:  # 源文件无表格，则返回1
        book = Workbook()  # 先创建一个工作簿
        del book["Sheet"]
        for s, tb in enumerate(row_content[:]):  # 读每个表数据
            sheet = book.create_sheet('Sheet' + str(s))  # 创建一个test_case的sheet表单
            for i, row in enumerate(tb[:]):  # 读每行数据
                for j, cell in enumerate(row[:]):  # 读每个单元格数据
                    sheet.cell(i + 1, j + 1, cell)
            sleep(1)
        book.save(dfoutname)
    else:
        print(f"{dfoutname}不含表格，未输出结果")


def pdf2office():  # 4智能转换 pdf转office
    cv = Converter(左表路径)
    cv.convert(os.path.join(D0R, OutN4ME).replace(".txt", ".docx"), start=0, end=None)
    cv.close()
    pdf = pdfplumber.open(左表路径)
    page = len(pdf.pages)
    print('总共有', page, '页')
    dfoutname = os.path.join(D0R, OutN4ME).replace(".txt", ".xlsx")
    writer = pd.ExcelWriter(dfoutname, engine='xlsxwriter')
    for i in range(0, page):
        print('正在输出第', str(i + 1), '页表格')
        p0 = pdf.pages[i]
        try:
            table = p0.extract_table()
            df = pd.DataFrame(table[1:], columns=table[0])
            df.to_excel(writer, sheet_name='Sheet' + str(i), index=False)
        except Exception as e:
            print(e)
            pass
    writer.close()
    pdf.close()
    if os.path.getsize(dfoutname) < 5555:
        os.remove(dfoutname)


def ofd2txt():  # 4智能转换 ofd转txt
    pages = []
    with ZipFile(左表路径, 'r') as f:
        for file in f.namelist():
            f.extract(file, T3MP)
    for filedir, dirs, files in os.walk(T3MP + "Doc_0\\Pages"):
        for filename in files:
            print(filedir + "\\" + filename)
            if "Page_" in filedir:
                pages.append(filedir)
    pages.sort(key=lambda x: int(x[x.rfind("_"):].replace("_", "")))
    total = ""
    for page in pages:
        xmlpath = page + "\\Content.xml"
        with open(xmlpath, 'r', encoding="utf-8") as openfile:
            text = openfile.read()
            text = re.sub(r"<.*?>", "", text)
            text = re.sub(r"M \d.*?\d L 0 0 C", "", text)
            total += text
    if total:
        dfoutname = os.path.join(D0R, OutN4ME)
        with open(dfoutname, 'w', encoding="utf-8") as openfile:
            openfile.writelines(total)


def split_df():  # 2筛选拆分计算 a拆分指定列@ 指定单列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            print("表格中不存在指定的列名，请核实")
            return
        groups = df.groupby(df[C0L])  # 仅支持单列
        if not os.path.exists(D0R):
            os.mkdir(D0R)
        outdir = os.path.join(D0R, os.path.basename(左表路径))
        outdir = outdir + '拆分结果'
        if not os.path.exists(outdir):
            os.mkdir(outdir)
        for group in groups:
            if (len(group[1]) >= 1048576) or 强制输出CSV:
                if group[0] == "":
                    group[1].to_csv(outdir + '\\空值.csv', index=False)
                else:
                    group[1].to_csv(outdir + "\\" + str(group[0]) + '.csv', index=False)
            else:
                if group[0] == "":
                    group[1].to_excel(outdir + '\\空值.xlsx', index=False)
                else:
                    group[1].to_excel(outdir + "\\" + str(group[0]) + '.xlsx', index=False)


def select_column(task):  # 2筛选拆分计算 g选取或删除列@@ 指定多列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not find_columns(df, C0L.replace("~", "").split(',')):
        print("表格中不存在指定的列名，请核实")
        return
    if not df.empty:
        if task == "删除":  # 删除指定列
            col = C0L.replace("~", "")
            siftdf = df.drop(col.split(','), axis=1)
        else:  # 选取指定列
            siftdf = df[C0L.split(',')]
        dftofile(siftdf)


def select_query_eval(task):  # 2筛选拆分计算 h自定query公式筛选 指定公式.txt
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if task == 'query':
            print(f"当前query公式：{参数}")
            siftdf = df.query(参数)
            dftofile(siftdf)
            return
        elif task == 'eval':
            print(f"当前eval公式：{参数}")
            siftdf = df.eval(参数)
            dftofile(siftdf)


def select_text(contain):  # 2筛选拆分计算 d筛选指定文本@# 指定单列+参数
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if not find_columns(df, C0L.split("|")):
            print("表格中不存在指定的列名，请核实")
            return
        if "str" not in str(type(df[C0L][0])):
            df[C0L] = df[C0L].astype("str")
        range_str = 参数.replace("-", "|")
        if contain:
            siftdf = df.loc[df[C0L].str.contains(range_str)]
        else:
            siftdf = df.loc[~df[C0L].str.contains(range_str)]
        dftofile(siftdf)
    else:
        print("指定的文件不含表格，请核实")


def del_duplicates():  # 2筛选拆分计算 2e删除重复值@@# 指定多列+参数
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        col = C0L.split(',')
        if not find_columns(df, col):
            print("表格中不存在指定的列名，请核实")
            return
        if 参数 == "0":
            df.drop_duplicates(subset=col, keep=False, inplace=True)
            print("False： 删除所有重复值")
        elif 参数 == "1":
            df.drop_duplicates(subset=col, keep="last", inplace=True)
            print("last： 保留最后一次出现的重复值")
        else:
            df.drop_duplicates(subset=col, keep="first", inplace=True)
            print("first： 保留第一次出现的重复值")
        dftofile(df)


def select_duplicates():  # 2筛选拆分计算 2f选取重复值@@# 指定多列+参数
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        col = C0L.split(',')
        if not find_columns(df, col):
            print("表格中不存在指定的列名，请核实")
            return
        if 参数 == "0":
            df = df[df.duplicated(subset=col, keep="first")]
            print("first： 除了第一次出现外，其余相同的被标记为重复")
        elif 参数 == "1":
            df = df[df.duplicated(subset=col, keep="last")]
            print("last：除了最后一次出现外，其余相同的被标记为重复")
        else:
            df = df[df.duplicated(subset=col, keep=False)]
            print("False：即所有相同的都被标记为重复")
        dftofile(df)


def append_df():  # 1匹配合并 f多表对齐同名列_纵向堆叠
    df = pd.DataFrame()
    for path in 左表路径:
        leftdf = file_read(os.path.dirname(path), os.path.basename(path))
        if not leftdf.empty:
            df = df.append(leftdf)
    dftofile(df)


def merge_df(h0w, notboth=False, db=False):  # 1匹配合并 两表横向@@ 指定多列
    global 左表工作簿, dfBIG, dfBIGc0l
    c0lleft = C0L.split(',')
    dfleft = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if dfleft.empty:
        print(f"左表是空表，终止执行")
        return

    左表工作簿 = 右表工作簿
    if db:
        if 右表路径:
            dfBIG = file_read(os.path.dirname(右表路径), os.path.basename(右表路径))
        if 右表列名:
            dfBIGc0l = 右表列名.split(',')
        else:
            print("未指定大表列名，终止执行")
    else:
        c0lright = 右表列名.split(',')
        dfright = file_read(os.path.dirname(右表路径), os.path.basename(右表路径))

    if '_本行来自' in dfleft.columns:
        dfleft = dfleft.drop('_本行来自', axis=1)
    if not find_columns(dfleft, c0lleft):
        print("左表或小表中不存在指定的列名，请核实")
        return

    if db:
        print(f"小表列名：【{c0lleft}】 大表列名：【{dfBIGc0l}】")
        if dfBIG.empty:
            print(f"大表是空表，终止执行")
            return
        if not find_columns(dfBIG, dfBIGc0l):
            print("大表中不存在指定的列名，请核实")
            return
        df = pd.merge(dfleft, dfBIG, how=h0w, left_on=c0lleft, right_on=dfBIGc0l, suffixes=('_来自小表', '_来自大表'),
                      indicator='_本行来自')
        df['_本行来自'].replace({'both': '两表', 'left_only': '小表', 'right_only': '大表'}, inplace=True)
        if df.shape[0] == 0:
            print(f"{'*' * 10} 小表和大表未匹配到任何内容，不写入文件 {'*' * 10}")
        else:
            dftofile(df)
    else:
        print(f"左表列名：【{c0lleft}】 右表列名：【{c0lright}】")
        if dfright.empty:
            print(f"右表是空表，终止执行")
            return
        if '_本行来自' in dfright.columns:
            dfright = dfright.drop('_本行来自', axis=1)
        if not find_columns(dfright, c0lright):
            print("右表中不存在指定的列名，请核实")
            return
        df = pd.merge(dfleft, dfright, how=h0w, left_on=c0lleft, right_on=c0lright, suffixes=('_来自左表', '_来自右表'),
                      indicator='_本行来自')
        df['_本行来自'].replace({'both': '两表', 'left_only': '左表', 'right_only': '右表'}, inplace=True)
        if notboth:
            df = df[df['_本行来自'] != '两表']
        dftofile(df)


def find_columns(df, cols):  # 查找列名cols（列表类型）是否存在
    finds = 0
    for col in cols:
        find = False
        for dfcol in df.columns:
            if col == dfcol:
                find = True
        if find:
            finds += 1
    if finds == len(cols):
        return True
    else:
        return False


def excelpivottable():  # 3分析处理 a生成数据透视表@ 指定多列
    df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
    if not df.empty:
        if not find_columns(df, C0L.split(",")):
            print("表格中不存在指定的列名，请核实")
            return
        col = C0L.split(",")
        col2 = 参数.split(",")
        dfresult = df.groupby(col, as_index=False)[col2[0]].agg(len)
        dfresult.rename(columns={col2[0]: "_计数"}, inplace=True)
        if "2" not in 求值及预处理类型:
            dfresult = dfresult.drop("_计数", axis=1)
        求值及预处理类型字典 = {}
        if "1" in 求值及预处理类型:
            求值及预处理类型字典.update({np.sum: "_求和："})
        if "3" in 求值及预处理类型:
            求值及预处理类型字典.update({np.mean: "_平均值："})
        if "4" in 求值及预处理类型:
            求值及预处理类型字典.update({max: "_最大值："})
        if "5" in 求值及预处理类型:
            求值及预处理类型字典.update({min: "_最小值："})
        for key, value in 求值及预处理类型字典.items():
            dfpivot = df.groupby(col, as_index=False)[col2].agg(key)
            dfpivot = dfpivot.drop(col, axis=1)
            for i in col2:
                dfpivot.rename(columns={i: value + i}, inplace=True)
            dfresult = pd.concat([dfresult, dfpivot], axis=1)
        dftofile(dfresult)


def dftofile(df):  # df写入文件 Excel使用zip64
    if not os.path.exists(D0R):
        os.mkdir(D0R)
    print("写入文件总行数：" + str(df.shape[0]))
    if (df.shape[0] >= 1048576) or 强制输出CSV:
        dfoutname = os.path.join(D0R, OutN4ME).replace(".txt", ".csv")
        df.to_csv(dfoutname, index=False)
    else:
        dfoutname = os.path.join(D0R, OutN4ME).replace(".txt", ".xlsx")
        writer = pd.ExcelWriter(dfoutname, engine='xlsxwriter')
        df.to_excel(writer, index=False, header=True)
        writer.book.use_zip64()
        writer.save()
    print("【写入文件】成功：" + dfoutname)


@get_time("读取文件")
def file_read(filedir, filename, transform=True):  # 根据文件.扩展名读取文件，transform默认为真：要做预处理
    filenameui = filename
    print(filenameui)
    filepath = os.path.join(filedir, filename)  # 读取文件路径
    if "." in filename:
        fileext = filename[filename.rfind("."):]  # 文件.扩展名
    else:
        fileext = ""
    filename = filename.replace(fileext, "")  # 去掉文件.扩展名
    fileext = fileext.lower()
    df = pd.DataFrame()
    if os.path.getsize(filepath):  # 文件大小不等于0才执行
        if ".xls" in fileext:
            if 左表工作簿:
                if 左表工作簿.isnumeric():
                    sheetname = int(左表工作簿) - 1
                else:
                    sheetname = 左表工作簿
            else:
                sheetname = 0
            df = pd.read_excel(filepath, keep_default_na=False, sheet_name=sheetname)
        else:
            # 判断csv文件编码及分隔符
            with open(filepath, 'rb') as openfile:
                enc0des = detect(openfile.readline())
                print("推测文件编码：", enc0des)
                enc0de = enc0des['encoding']
                if "utf" not in enc0de.lower():
                    enc0de = "GB18030"
                # if enc0de == ("GB2312" or "ISO" or "TIS-620" in enc0de) or (enc0de is None):
                #    enc0de = "GB18030"
                print("确定文件编码：", enc0de)
            with open(filepath, 'r', encoding=enc0de) as openfile:
                text3 = openfile.readline() + openfile.readline() + openfile.readline()
                print("文件前3行内容：", text3)
                text3 = re.sub("[^,;|.\t]", "", text3)  # 选取分隔符：【,】【;】【|】【.】【   】
                print("文件前3行分隔符：", text3)
                textsep = textmax(text3)
                print("判定分隔符为：【" + textsep + "】")
            # 处理跳行警告
            with open("Log.txt", 'w') as openfile:
                with redirect_stderr(openfile):
                    df = pd.read_csv(filepath, encoding=enc0de, sep=textsep, engine='python', error_bad_lines=False,
                                     keep_default_na=False)  # skipinitialspace =  忽略分隔符后的空白
            with open("Log.txt", 'r') as openfile:
                textlist = openfile.readlines()
                filetext = []
                for text in textlist:
                    if "Skipping line" in text:
                        text = text.replace("Skipping line", "最终结果不含原文件的第")
                        text = text.replace(":", "行，请检查，具体原因：")
                        filetext.append(text)
            if filetext:
                if not os.path.exists(D0R):
                    os.mkdir(D0R)
                with open(f"{D0R}\\跳行警告{T1ME}：{filenameui}.txt", 'w') as openfile:
                    openfile.writelines(filetext)
            if os.path.exists("Log.txt"):
                os.remove("Log.txt")
                sleep(1)
        print(f"【读取文件】成功：{filename}，总行数{df.shape[0]}")
        if transform:
            if 删除空格:
                # 所有文本列去除空格
                list_str = list(df.select_dtypes(include=["object"]))
                for i in list_str:
                    df[i].replace(" ", "", regex=True, inplace=True)  # 正则去除所有空格
            if 删除制表符:
                # 所有文本列去除TAB空白
                list_str = list(df.select_dtypes(include=["object"]))
                for i in list_str:
                    df[i].replace("\t", "", regex=True, inplace=True)  # 正则去除所有TAB空白
            if 数值类型最大位数 != 0:
                # 所有数值列11位及以下转为文本
                list_num = list(df.select_dtypes(include=["int", "int64", "uint64", "float", "float64"]))
                print("所有类型为数值的列：", list_num)
                for i in list_num:
                    # print(max(df[i].astype("str").apply(len)))  # 每列最大位数
                    if not max(df[i].astype("str").apply(len)) <= int(数值类型最大位数):
                        df[i] = df[i].astype("str")
                list_num = list(df.select_dtypes(include=["int", "int64", "uint64", "float", "float64"]))
                print(f"{数值类型最大位数}位及以下的数值列：{list_num}")
    else:
        print(f"{filename}文件为空，终止执行")
        return df
    return df


def textmax(a):  # 计算字符串出现最多的符号
    t = ""
    k = 0
    y = []
    for q in a:
        s = a.count(q)
        y.append(s)
        t = max(y)
    for i in range(len(a)):
        if t != y[i]:
            continue
        k = a[i]
    print("出现最多的符号：【" + k + "】，" + str(t) + "次")
    return k


def run_python():
    dfcount = 0
    for path in 左表路径:
        dfcount += 1
        exec(f"df{dfcount} = file_read(os.path.dirname(path), os.path.basename(path))")
        print(f"已经读取表格{path}到 df{dfcount}")
    df = pd.DataFrame()
    with open(D0R + "python语句.txt", 'r', encoding="GB2312") as openfile:
        runpython = openfile.read()
    exec(runpython)
    if "ex = True" in runpython:
        runpython = ""
        while True:
            print("输入python语句回车后存储（直接回车开始执行所有语句，输入end回车结束操作）")
            inputtxt = input("请输入：")
            if inputtxt == "":
                exec(runpython)
                print("已执行存储的语句并清空：")
                print(runpython)
                runpython = ""
            elif inputtxt == "end":
                break
            else:
                runpython += (inputtxt + "\n")
    dftofile(df)


def scantask(scandir):  # 扫描任务
    global Fi1eEXT, C0L, 右表列名, Fi1eTASK, T1ME, 左表路径, 右表路径, OutN4ME, 参数, 求值及预处理类型, 左表工作簿, 右表工作簿,\
        强制输出CSV, 删除空格, 删除制表符, 数值类型最大位数, 任务类型
    for filedir, dirs, files in os.walk(scandir):
        for Fi1eTASK in files:
            if ".txt" in Fi1eTASK.lower():
                taskpath = os.path.join(filedir, Fi1eTASK)
                with open(taskpath, 'r', encoding="GB2312") as openfile:
                    tasks = openfile.readlines()
                os.remove(taskpath)
                for task in tasks:
                    task = eval(task)
                    T1ME = datetime.datetime.now().strftime('%H%M%S')
                    C0L = task["左表列名"]  # task["左表列名"]
                    右表列名 = task["右表列名"]  # task["右表列名"]
                    任务类型 = task["任务类型"]  # task["任务类型"]
                    左表路径 = task["左表"]  # task["左表"]
                    右表路径 = task["右表"]  # task["右表"]
                    参数 = task["参数"]  # task["参数"]
                    求值及预处理类型 = task["求值及预处理类型"]  # task["求值及预处理类型"]
                    强制输出CSV = True if "6" in 求值及预处理类型 else False
                    删除空格 = True if "7" in 求值及预处理类型 else False
                    删除制表符 = True if "8" in 求值及预处理类型 else False
                        
                    数值类型最大位数 = int(task["数值类型最大位数"])  # int(task["数值类型最大位数"])
                    左表工作簿 = task["左表工作簿"]  # task["左表工作簿"]
                    右表工作簿 = task["右表工作簿"]  # task["右表工作簿"]
                    if not os.path.exists(D0R):
                        os.mkdir(D0R)
                    try:
                        start_times = time()
                        if 任务类型 == "1f" or 任务类型 == '2h':
                            OutN4ME = os.path.basename(左表路径[0]) + Fi1eTASK
                        else:
                            OutN4ME = os.path.basename(左表路径) + Fi1eTASK
                            if "." in os.path.basename(左表路径):
                                Fi1eEXT = 左表路径[左表路径.rfind("."):]  # 文件.扩展名
                            else:
                                Fi1eEXT = ""
                        if 任务类型 == "1a":
                            print("【执行】1匹配合并 a两表横向并集_输出全部@")
                            merge_df("outer")
                        elif 任务类型 == "1b":
                            print("【执行】1匹配合并 b两表横向交集_输出相同@")
                            merge_df("inner")
                        elif 任务类型 == "1c":
                            print("【执行】1匹配合并 c_两表横向合并_反集")
                            merge_df("outer", True)
                        elif 任务类型 == "1d":
                            print("【执行】1匹配合并 d两表横向左集_输出左表@")
                            merge_df("left")
                        elif 任务类型 == "1e":
                            print("【执行】1匹配合并 e两表横向右集_输出右表@")
                            merge_df("right")
                        elif 任务类型 == "1f":
                            print("【执行】1匹配合并 f多表对齐同名列_纵向堆叠")
                            append_df()
                        elif 任务类型 == "2a":
                            print("【执行】2筛选拆分计算 a拆分指定列@")
                            split_df()
                        elif 任务类型 == "2b":
                            print("【执行】2筛选拆分计算 b自定query公式筛选")
                            select_query_eval('query')
                        elif 任务类型 == "2c":
                            print("【执行】2筛选拆分计算 c自定eval公式计算")
                            select_query_eval('eval')
                        elif 任务类型 == "2d":
                            print("【执行】2筛选拆分计算 d删除重复值@@#")
                            del_duplicates()
                        elif 任务类型 == "2e":
                            print("【执行】2筛选拆分计算 e选取重复值@@#")
                            select_duplicates()
                        elif 任务类型 == "2f":
                            print("【执行】2筛选拆分计算 f删除指定列@@")
                            select_column("删除")
                        elif 任务类型 == "2g":
                            print("【执行】2筛选拆分计算 g选取指定列@@")
                            select_column("选取")
                        elif 任务类型 == "2h":
                            print("【执行】2h自定python语句操作")
                            run_python()
                        elif 任务类型 == "2i":
                            print("【执行】2i筛选包含文本@#")
                            select_text(True)
                        elif 任务类型 == "2j":
                            print("【执行】2j筛选不包含文本@#")
                            select_text(False)
                        elif 任务类型 == "3a":
                            print("【执行】3分析处理 a生成数据透视表@@")
                            excelpivottable()
                        elif 任务类型 == "3b":
                            print("【执行】3分析处理 b分析身份证@")
                            idcard_analyse()
                        elif 任务类型 == "3c":
                            print("【执行】3分析处理 c分析机构号@")
                            network_analyse("_机构号")
                        elif 任务类型 == "3d":
                            print("【执行】3分析处理 d分析网点名@")
                            network_analyse("_精简名拼音")
                        elif 任务类型 == "3e":
                            print("【执行】3分析处理 e分析表格结构@@")
                            excel_analyse()
                        elif 任务类型 == "3f":  # 未完成
                            print("【执行】3分析处理 f获取指定列拼音@")
                            get_pinyin()
                        elif 任务类型 == "3g":  # 未完成
                            print("【执行】3分析处理 g八位列转日期格式")
                            get_date()
                        elif 任务类型 == "4a":
                            print("【执行】4智能转换")
                            if "pdf" in Fi1eEXT:
                                pdf2office()
                            elif "docx" in Fi1eEXT:
                                word2excel()
                            elif "doc" in Fi1eEXT:
                                print(f"不支持doc文件{os.path.basename(左表路径)}，请先转换为docx")
                            elif ".ofd" in Fi1eEXT:
                                ofd2txt()
                            else:
                                df = file_read(os.path.dirname(左表路径), os.path.basename(左表路径))
                                if not df.empty:
                                    dftofile(df)
                                else:
                                    print("表格为空，未生成文件")
                        elif 任务类型 == "4b":
                            print("【执行】4b大小表连续比对，输出小表@@")
                            merge_df("left", False, True)
                        elif 任务类型 == "4c":
                            print("【执行】4b大小表连续比对，输出交集@@")
                            merge_df("inner", False, True)
                        end_times = time()
                        print("【执行成功】")
                        print(f'任务总耗时：{end_times - start_times}秒')
                    except Exception as e:
                        filepath = f"{D0R}\\失败原因：{OutN4ME}"
                        with open(filepath, 'w') as openfile:
                            openfile.write(str(e))
                        print(e, "【执行失败】见失败原因：txt")
                    finally:
                        print("────────────────────────────────────────────────────────────────────────────────────")


T3MP = os.getcwd() + "\\TEMP\\"  # 临时目录
D0R = os.getcwd() + "\\结果\\"  # 结果根目录
D1R = os.getcwd() + "\\任务\\"  # 任务根目录
Fi1eTASK = ""  # 任务文件名.txt
Fi1eEXT = ""  # 左表文件扩展名
T1ME = ""  # 时分秒
数值类型最大位数 = 0
强制输出CSV = 删除空格 = 删除制表符 = False
C0L = 右表列名 = 左表路径 = 右表路径 = OutN4ME = 参数 = 任务类型 = 求值及预处理类型 = 左表工作簿 = 右表工作簿 = 八位日期列 = ""
IDcardAERA = pd.DataFrame()  # _机构号网点名_身份证代码地区_转换表.xlsx sheet_name='身份证代码to地区'
dfBIG = pd.DataFrame()  # 大小表连续比对中的大表
dfBIGc0l = ""  # 大小表连续比对中的大表列名
if __name__ == "__main__":
    print(f"{'-'*10} 程序运行中，请生成任务 {'-'*10}")
    if not os.path.exists(D1R):
        os.mkdir(D1R)
    if not os.path.exists(D0R):
        os.mkdir(D0R)
    while True:
        sleep(1)
        if os.path.exists(T3MP):
            rmtree(T3MP)  # 删除临时目录
        try:
            scantask(D1R)
        except RuntimeError as ec:
            pass
